"""
src/exporter.py – Report Generator (Phase 4 + Phase 5)
=======================================================
Exports the ``MonthlyBrandSummary`` DataFrame to a multi-tab Excel
workbook with one ``{Brand} Financial`` sheet per brand, and optionally
a ``Summary Campaigns`` tab with campaign KPI metrics.

Uses openpyxl for cell-level formatting:
  • Percentage columns → 0.00 % format
  • Currency / GGR columns → #,##0.00 number format
  • Header row styling (bold, coloured background)

Spec refs: §3-C, §2 MonthlyBrandSummary, §2 CampaignSummary, §2 CohortMatrix, §4.
"""
from __future__ import annotations

import calendar
import io
import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Output defaults ──────────────────────────────────────────────────────
OUTPUT_FILENAME = "Summary_Data_Auto.xlsx"

# Column display names for the sheet header row
COLUMN_HEADERS: list[str] = [
    "Month",
    "Brand",
    "Negative Yield Players",
    "Profitable Players",
    "Flat",
    "Total Players",
    "Profitable %",
    "GGR",
    "Total Handle",
    "Hold %",
    "GGR Per Player",
    "Top 10% GGR Share",
    "New Players",
    "Returning Players",
    "Retention %",
    "NGR",
    "Turnover (Casino)",
    "GGR (Casino)",
    "NGR (Casino)",
    "Turnover (Sports)",
    "GGR (Sports)",
    "NGR (Sports)",
]

# Internal DataFrame column names (same order as COLUMN_HEADERS)
DF_COLS: list[str] = [
    "month",
    "brand",
    "negative_yield_players",
    "profitable_players",
    "flat",
    "total_players",
    "profitable_pct",
    "ggr",
    "total_handle",
    "hold_pct",
    "ggr_per_player",
    "top_10_pct_ggr_share",
    "new_players",
    "returning_players",
    "retention_pct",
    "ngr",
    "turnover_casino",
    "ggr_casino",
    "ngr_casino",
    "turnover_sports",
    "ggr_sports",
    "ngr_sports",
]

# Formatting specs  (0-based column index → openpyxl number format)
_PCT_FMT = "0.00%"
_NUM_FMT = "#,##0.00"
_INT_FMT = "#,##0"

# Column indices that need special formatting (0-based)
_FORMAT_MAP: dict[int, str] = {
    2:  _INT_FMT,   # negative_yield_players
    3:  _INT_FMT,   # profitable_players
    4:  _INT_FMT,   # flat
    5:  _INT_FMT,   # total_players
    6:  _PCT_FMT,   # profitable_pct
    7:  _NUM_FMT,   # ggr
    8:  _NUM_FMT,   # total_handle
    9:  _PCT_FMT,   # hold_pct
    10: _NUM_FMT,   # ggr_per_player
    11: _PCT_FMT,   # top_10_pct_ggr_share
    12: _INT_FMT,   # new_players
    13: _INT_FMT,   # returning_players
    14: _PCT_FMT,   # retention_pct
    15: _NUM_FMT,   # ngr
    16: _NUM_FMT,   # turnover_casino
    17: _NUM_FMT,   # ggr_casino
    18: _NUM_FMT,   # ngr_casino
    19: _NUM_FMT,   # turnover_sports
    20: _NUM_FMT,   # ggr_sports
    21: _NUM_FMT,   # ngr_sports
}

# ── Campaign tab config ──────────────────────────────────────────────────
CAMPAIGN_HEADERS: list[str] = [
    "Month",
    "Brand",
    "Records",
    "KPI 1 - Conversions",
    "KPI 2 - Logins",
    "Calls",
    "Emails Sent",
    "SMS Sent",
    "KPI1 Conversion Rate",
    "KPI2 Login Rate",
]

CAMPAIGN_DF_COLS: list[str] = [
    "month",
    "brand",
    "total_records",
    "total_kpi1",
    "total_kpi2",
    "total_calls",
    "total_emails",
    "total_sms",
    "kpi1_conversion_rate",
    "kpi2_login_rate",
]

_CAMPAIGN_FORMAT_MAP: dict[int, str] = {
    2: _INT_FMT,  # total_records
    3: _INT_FMT,  # total_kpi1
    4: _INT_FMT,  # total_kpi2
    5: _INT_FMT,  # total_calls
    6: _INT_FMT,  # total_emails
    7: _INT_FMT,  # total_sms
    8: _PCT_FMT,  # kpi1_conversion_rate
    9: _PCT_FMT,  # kpi2_login_rate
}

# ── Segmentation tab config ──────────────────────────────────────────────
SEGMENTATION_HEADERS: list[str] = [
    "Month",
    "Brand",
    "Segment (WB Tag)",
    "Total Players",
    "GGR",
]

SEGMENTATION_DF_COLS: list[str] = [
    "month",
    "brand",
    "wb_tag",
    "total_players",
    "ggr",
]

_SEGMENTATION_FORMAT_MAP: dict[int, str] = {
    3: _INT_FMT,  # total_players
    4: _NUM_FMT,  # ggr
}

# ── Both Business tab config (Phase 9) ───────────────────────────────────
BOTH_BUSINESS_HEADERS: list[str] = [
    "Month",
    "Turnover",
    "GGR",
    "Margin %",
    "Rev Share (15%)",
    "Net Income",
    "New Players",
    "Returning Players",
    "Reactivated Players",
    "Conversions",
    "Total Players",
    "Profitable Players",
    "Neg. Yield Players",
    "New Players %",
    "Returning Players %",
    "GGR / Player",
    "Turnover / Player",
    "Income / Player",
    "New Player GGR",
    "Returning Player GGR",
    "NGR",
    "Turnover (Casino)",
    "GGR (Casino)",
    "NGR (Casino)",
    "Turnover (Sports)",
    "GGR (Sports)",
    "NGR (Sports)",
]

BOTH_BUSINESS_DF_COLS: list[str] = [
    "month",
    "turnover",
    "ggr",
    "margin",
    "revenue_share_deduction",
    "net_income",
    "new_players",
    "returning_players",
    "reactivated_players",
    "conversions",
    "total_players",
    "profitable_players",
    "negative_yield_players",
    "new_players_pct",
    "returning_players_pct",
    "ggr_per_player",
    "turnover_per_player",
    "income_per_player",
    "new_player_ggr",
    "returning_player_ggr",
    "ngr",
    "turnover_casino",
    "ggr_casino",
    "ngr_casino",
    "turnover_sports",
    "ggr_sports",
    "ngr_sports",
]

_BOTH_BUSINESS_FORMAT_MAP: dict[int, str] = {
    1: _NUM_FMT,   # turnover
    2: _NUM_FMT,   # ggr
    3: _PCT_FMT,   # margin
    4: _NUM_FMT,   # revenue_share_deduction
    5: _NUM_FMT,   # net_income
    6: _INT_FMT,   # new_players
    7: _INT_FMT,   # returning_players
    8: _INT_FMT,   # reactivated_players
    9: _INT_FMT,   # conversions
    10: _INT_FMT,  # total_players
    11: _INT_FMT,  # profitable_players
    12: _INT_FMT,  # negative_yield_players
    13: _PCT_FMT,  # new_players_pct
    14: _PCT_FMT,  # returning_players_pct
    15: _NUM_FMT,  # ggr_per_player
    16: _NUM_FMT,  # turnover_per_player
    17: _NUM_FMT,  # income_per_player
    18: _NUM_FMT,  # new_player_ggr
    19: _NUM_FMT,  # returning_player_ggr
    20: _NUM_FMT,  # ngr
    21: _NUM_FMT,  # turnover_casino
    22: _NUM_FMT,  # ggr_casino
    23: _NUM_FMT,  # ngr_casino
    24: _NUM_FMT,  # turnover_sports
    25: _NUM_FMT,  # ggr_sports
    26: _NUM_FMT,  # ngr_sports
}

# Header styling
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Cohort matrix header styling (distinct green)
_COHORT_HEADER_FILL = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
_COHORT_TITLE_FONT = Font(bold=True, size=12, color="375623")

# Both Business header styling (gold)
_BB_HEADER_FILL = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")


# ═══════════════════════════════════════════════════════════════════════════
#  Public API
# ═══════════════════════════════════════════════════════════════════════════
def export_to_excel(
    summary_df: pd.DataFrame,
    campaign_df: pd.DataFrame | None = None,
    cohort_matrices: dict[str, pd.DataFrame] | None = None,
    segmentation_df: pd.DataFrame | None = None,
    both_business_df: pd.DataFrame | None = None,
    ops_df: pd.DataFrame | None = None,
) -> io.BytesIO:
    """Write *summary_df* to a multi-tab Excel workbook in-memory.

    One sheet per brand, named ``{Brand} Financial``.  If *campaign_df* is
    provided and non-empty, a ``Summary Campaigns`` tab is also written.
    If *cohort_matrices* is provided, each brand's cohort retention
    matrix is appended below its financial summary.

    Returns
    -------
    io.BytesIO
        In-memory buffer containing the Excel workbook.
    """
    buf = io.BytesIO()

    brands = sorted(summary_df["brand"].unique())

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ── Both Business Summary tab (Phase 9 — first tab) ──────────
        if both_business_df is not None and not both_business_df.empty:
            _write_both_business_tab(writer, both_business_df)
            logger.info("Exported 'Both Business Summary' tab (%d rows)", len(both_business_df))

        # ── Financial tabs (one per brand) ────────────────────────────
        for brand in brands:
            sheet_name = f"{brand} Financial"
            brand_df = (
                summary_df[summary_df["brand"] == brand]
                .sort_values("month")
                .reset_index(drop=True)
            )

            display = _prepare_display_df(brand_df)

            display.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=1,
                header=False,
            )

            ws = writer.sheets[sheet_name]
            _write_header(ws, COLUMN_HEADERS)
            _apply_formatting(ws, num_data_rows=len(display), fmt_map=_FORMAT_MAP)
            _auto_column_widths(ws, COLUMN_HEADERS)

            # ── Cohort matrix below financial data (Phase 7) ─────────
            if cohort_matrices and brand in cohort_matrices:
                cohort_df = cohort_matrices[brand]
                if not cohort_df.empty:
                    _write_cohort_section(
                        ws,
                        cohort_df,
                        start_row=len(display) + 4,  # 2 blank rows gap
                    )
                    logger.info("Wrote cohort matrix for %s (%d cohorts)", brand, len(cohort_df))

        # ── Campaign tab (Phase 5) ───────────────────────────────────
        if campaign_df is not None and not campaign_df.empty:
            _write_campaign_tab(writer, campaign_df)
            logger.info("Exported 'Summary Campaigns' tab (%d rows)", len(campaign_df))

        # ── Segmentation tab (Phase 8) ────────────────────────────────
        if segmentation_df is not None and not segmentation_df.empty:
            _write_segmentation_tab(writer, segmentation_df)
            logger.info("Exported 'Segmentation' tab (%d rows)", len(segmentation_df))

        # ── Operations tab (Phase 23) ────────────────────────────────
        if ops_df is not None and not ops_df.empty:
            _write_ops_tab(writer, ops_df)
            logger.info("Exported 'Operations Tracker' tab (%d rows)", len(ops_df))

        logger.info("Exported %d brand(s) to in-memory buffer", len(brands))

    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════
#  Internal helpers
# ═══════════════════════════════════════════════════════════════════════════
def _prepare_display_df(brand_df: pd.DataFrame) -> pd.DataFrame:
    """Build a display-ready copy with human-readable month names
    and percentage values divided by 100 (so openpyxl % format works).
    """
    out = brand_df[DF_COLS].copy()

    # Convert "YYYY-MM" → "Month YYYY" (e.g. "2024-08" → "August 2024")
    out["month"] = out["month"].apply(_pretty_month)

    # openpyxl percentage format expects 0–1 range, not 0–100
    for pct_col in ("profitable_pct", "hold_pct", "top_10_pct_ggr_share", "retention_pct"):
        if pct_col in out.columns:
            out[pct_col] = out[pct_col] / 100.0

    return out


def _write_cohort_section(ws, cohort_df: pd.DataFrame, start_row: int) -> None:  # noqa: ANN001
    """Write a cohort retention matrix below the financial data.

    Layout (1-indexed rows, starting at *start_row*):
      Row start_row:      Title  "Cohort Retention Matrix (%)"
      Row start_row + 1:  Header row (Acquisition Month | Month 1 | Month 2 | …)
      Row start_row + 2…: Data rows
    """
    # ── Title row ────────────────────────────────────────────────────────
    title_cell = ws.cell(row=start_row, column=1, value="Cohort Retention Matrix (%)")
    title_cell.font = _COHORT_TITLE_FONT

    # ── Header row ───────────────────────────────────────────────────────
    header_row = start_row + 1
    headers = ["Acquisition Month"] + list(cohort_df.columns)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _COHORT_HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # ── Data rows ────────────────────────────────────────────────────────
    for row_offset, (acq_month, row_data) in enumerate(cohort_df.iterrows()):
        data_row = header_row + 1 + row_offset

        # Acquisition month label (human-readable)
        acq_cell = ws.cell(row=data_row, column=1, value=_pretty_month(str(acq_month)))
        acq_cell.font = Font(bold=True)

        # Retention percentages
        for col_offset, value in enumerate(row_data):
            cell = ws.cell(row=data_row, column=2 + col_offset)
            if pd.notna(value):
                cell.value = value / 100.0  # openpyxl % format expects 0–1
                cell.number_format = _PCT_FMT
            cell.alignment = Alignment(horizontal="right")

    # Auto-width for cohort columns
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        current_width = ws.column_dimensions[col_letter].width or 0
        new_width = max(len(header) + 4, 14)
        ws.column_dimensions[col_letter].width = max(current_width, new_width)


def _write_campaign_tab(writer, campaign_df: pd.DataFrame) -> None:  # noqa: ANN001
    """Write the Summary Campaigns tab to the workbook."""
    sheet_name = "Summary Campaigns"

    display = campaign_df[CAMPAIGN_DF_COLS].copy()
    display = display.sort_values(["month", "brand"]).reset_index(drop=True)
    display["month"] = display["month"].apply(_pretty_month)

    # openpyxl percentage format expects 0–1 range
    for rate_col in ("kpi1_conversion_rate", "kpi2_login_rate"):
        if rate_col in display.columns:
            display[rate_col] = display[rate_col] / 100.0

    display.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        startrow=1,
        header=False,
    )

    ws = writer.sheets[sheet_name]
    _write_header(ws, CAMPAIGN_HEADERS)
    _apply_formatting(ws, num_data_rows=len(display), fmt_map=_CAMPAIGN_FORMAT_MAP)
    _auto_column_widths(ws, CAMPAIGN_HEADERS)


def _write_segmentation_tab(writer, seg_df: pd.DataFrame) -> None:  # noqa: ANN001
    """Write the Segmentation tab to the workbook."""
    sheet_name = "Segmentation"

    display = seg_df[SEGMENTATION_DF_COLS].copy()
    display = display.sort_values(["month", "brand", "wb_tag"]).reset_index(drop=True)
    display["month"] = display["month"].apply(_pretty_month)

    display.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        startrow=1,
        header=False,
    )

    ws = writer.sheets[sheet_name]
    _write_header(ws, SEGMENTATION_HEADERS)
    _apply_formatting(ws, num_data_rows=len(display), fmt_map=_SEGMENTATION_FORMAT_MAP)
    _auto_column_widths(ws, SEGMENTATION_HEADERS)


def _write_both_business_tab(writer, bb_df: pd.DataFrame) -> None:  # noqa: ANN001
    """Write the Both Business Summary tab to the workbook."""
    sheet_name = "Both Business Summary"

    display = bb_df[BOTH_BUSINESS_DF_COLS].copy()
    display = display.sort_values("month").reset_index(drop=True)
    display["month"] = display["month"].apply(_pretty_month)

    # openpyxl percentage format expects 0–1 range
    for pct_col in ("margin", "new_players_pct", "returning_players_pct"):
        if pct_col in display.columns:
            display[pct_col] = display[pct_col] / 100.0

    display.to_excel(
        writer,
        sheet_name=sheet_name,
        index=False,
        startrow=1,
        header=False,
    )

    ws = writer.sheets[sheet_name]
    # Gold header styling for the primary aggregation tab
    for col_idx, header in enumerate(BOTH_BUSINESS_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _BB_HEADER_FILL
        cell.alignment = _HEADER_ALIGN
    _apply_formatting(ws, num_data_rows=len(display), fmt_map=_BOTH_BUSINESS_FORMAT_MAP)
    _auto_column_widths(ws, BOTH_BUSINESS_HEADERS)


def _write_header(ws, headers: list[str]) -> None:  # noqa: ANN001
    """Write styled column headers to row 1."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _apply_formatting(  # noqa: ANN001
    ws, num_data_rows: int, fmt_map: dict[int, str]
) -> None:
    """Apply number formats to data cells (rows 2 … N+1)."""
    for row in range(2, num_data_rows + 2):
        for col_0, fmt in fmt_map.items():
            cell = ws.cell(row=row, column=col_0 + 1)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")


def _auto_column_widths(ws, headers: list[str]) -> None:  # noqa: ANN001
    """Set column widths based on header length (with a reasonable min)."""
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = max(len(header) + 4, 14)
        ws.column_dimensions[col_letter].width = width


def _pretty_month(ym: str) -> str:
    """Convert 'YYYY-MM' → 'Month YYYY'."""
    y, m = (int(x) for x in ym.split("-"))
    return f"{calendar.month_name[m]} {y}"

def export_ops_to_excel(ops_df: pd.DataFrame) -> io.BytesIO:
    """Standalone export for Operations data."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write_ops_tab(writer, ops_df)
    buf.seek(0)
    return buf

def _write_ops_tab(writer, ops_df: pd.DataFrame) -> None:
    """Write the Operations data to the workbook."""
    sheet_name = "Operations Tracker"
    # Ensure we don't hit Excel sheet name length limits
    ops_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1, header=False)
    ws = writer.sheets[sheet_name]
    headers = list(ops_df.columns)
    _write_header(ws, headers)
    _auto_column_widths(ws, headers)

